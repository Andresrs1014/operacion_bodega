"""
Router: catálogo de Unidades de Empaque (UE) por referencia.
Prefijo: /productos/ue
"""
import io
from typing import List, Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from sqlalchemy.orm import Session
import openpyxl

from database import get_db
from dependencies import get_current_user, require_supervisor_or_admin
from models import Usuario
from schemas import (
    ImportResult,
    ImportRowError,
    ProductoUeOut,
    ProductoUeCreate,
    ProductoUeUpdate,
    ProductoUePage,
    ProductoUeBatchRequest,
    ProductoUeBatchItem,
)
from services.productos import (
    list_productos_ue,
    get_producto_ue,
    create_producto_ue,
    update_producto_ue,
    delete_producto_ue,
    batch_productos_ue,
)

router = APIRouter(prefix="/productos/ue", tags=["Productos UE"])

# Aliases de columna aceptados (case-insensitive, strip aplicado antes de comparar)
_ALIAS_REF: frozenset[str] = frozenset({
    "referencia", "ref", "codigo", "código", "cod",
    "codigo referencia", "código referencia", "codigo producto", "code",
})
_ALIAS_UE: frozenset[str] = frozenset({
    "ue", "unidad_empaque", "unidad de empaque", "unidades de empaque",
    "unidad empaque", "und emp", "und. de empaque", "und. empaque",
    "cantidad empaque", "u.e.", "unidades", "und",
})


def _buscar_col(headers: list[str], aliases: frozenset[str]) -> int | None:
    """Detecta índice de columna: exact match primero, luego partial match.
    Evita falsos positivos usando longitud mínima de 2 chars en partial."""
    for i, h in enumerate(headers):
        if h in aliases:
            return i
    for i, h in enumerate(headers):
        if any(len(a) >= 2 and (a in h or h in a) for a in aliases):
            return i
    return None


@router.get("", response_model=ProductoUePage)
def listar_ue(
    q: Optional[str] = Query(None, description="Búsqueda por referencia o texto"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _: Usuario = Depends(require_supervisor_or_admin),
):
    return list_productos_ue(db, q, page, limit)


@router.post("", response_model=ProductoUeOut, status_code=status.HTTP_201_CREATED)
def crear_ue(
    body: ProductoUeCreate,
    db: Session = Depends(get_db),
    _: Usuario = Depends(require_supervisor_or_admin),
):
    return create_producto_ue(db, body.referencia, body.unidad_empaque, body.texto_unidad_empaque)


@router.patch("/{producto_id}", response_model=ProductoUeOut)
def editar_ue(
    producto_id: int,
    body: ProductoUeUpdate,
    db: Session = Depends(get_db),
    _: Usuario = Depends(require_supervisor_or_admin),
):
    return update_producto_ue(db, producto_id, body.unidad_empaque, body.texto_unidad_empaque)


@router.delete("/{producto_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_ue(
    producto_id: int,
    db: Session = Depends(get_db),
    _: Usuario = Depends(require_supervisor_or_admin),
):
    delete_producto_ue(db, producto_id)
    return None


@router.post("/batch", response_model=List[ProductoUeBatchItem])
def batch_ue(
    body: ProductoUeBatchRequest,
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    """Devuelve solo las referencias que existen en el catálogo. Usado por validación."""
    productos = batch_productos_ue(db, body.referencias)
    return [
        ProductoUeBatchItem(
            referencia=p.referencia,
            unidad_empaque=p.unidad_empaque,
            texto_unidad_empaque=p.texto_unidad_empaque,
        )
        for p in productos
    ]


@router.post("/import", response_model=ImportResult)
async def importar_ue(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: Usuario = Depends(require_supervisor_or_admin),
):
    """Import masivo desde Excel. Columnas requeridas: Referencia, UE."""
    MAX_SIZE = 10 * 1024 * 1024  # 10 MB
    MAX_ROWS = 10_000

    contents = await file.read()
    if len(contents) > MAX_SIZE:
        raise HTTPException(status_code=413, detail="Archivo demasiado grande (máx 10 MB).")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel.")

    # Detectar encabezados en fila 1 (case-insensitive, trim)
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_ref = _buscar_col(headers, _ALIAS_REF)
    idx_ue  = _buscar_col(headers, _ALIAS_UE)
    if idx_ref is None or idx_ue is None:
        faltantes = []
        if idx_ref is None: faltantes.append("Referencia")
        if idx_ue  is None: faltantes.append("Unidad de Empaque")
        raise HTTPException(
            status_code=422,
            detail=f"Columnas no encontradas: {', '.join(faltantes)}. "
                   f"Encabezados detectados: {headers or ['(vacío)']}",
        )

    # Leer filas de datos
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(rows) >= MAX_ROWS:
            break
        ref_val = row[idx_ref] if idx_ref < len(row) else None
        ue_val  = row[idx_ue]  if idx_ue  < len(row) else None
        if ref_val is None and ue_val is None:
            continue  # fila vacía
        rows.append({"referencia": ref_val, "unidad_empaque": ue_val})

    if not rows:
        raise HTTPException(status_code=422, detail="El archivo no contiene filas de datos.")

    from services.productos import import_productos_ue
    result = import_productos_ue(db, rows)
    return ImportResult(
        ok=result["ok"],
        creadas=result["creadas"],
        actualizadas=result["actualizadas"],
        errores=[ImportRowError(**e) for e in result["errores"]],
    )


@router.get("/plantilla")
def descargar_plantilla(_: Usuario = Depends(require_supervisor_or_admin)):
    """Descarga una plantilla Excel con columnas Referencia, Unidad de Empaque y Texto."""
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UE"

    # Encabezados
    ws.append(["Referencia", "Unidad de Empaque", "Texto"])
    for col_letter, bold in [("A", True), ("B", True), ("C", True)]:
        ws[f"{col_letter}1"].font = Font(bold=True)
        ws[f"{col_letter}1"].fill = PatternFill("solid", fgColor="E2E8F0")

    # Filas ejemplo
    ws.append(["EJEMPLO-001", 6, "PAR X 6"])
    ws.append(["EJEMPLO-002", 12, "DOCENA"])
    ws.append(["EJEMPLO-003", 1, "UNIDAD"])

    # Formato numérico entero en columna B para evitar auto-formato fecha de Excel
    for row_idx in range(2, 5002):  # primeras 5000 filas de datos
        ws[f"B{row_idx}"].number_format = "0"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_ue.xlsx"},
    )
